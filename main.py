"""
main.py
入口程式：讀取 ./prompts/prompts_use.xlsx，逐一送至 grok.com/imagine 生成影片，
下載至 ./output/<日期>/，完成整列後在 D 欄寫入 Y。
"""
import asyncio
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

from grok_automation import GrokVideoAutomation

# ──────────────────────────────────────────────
# 路徑設定
# ──────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
PROMPTS_DIR = BASE_DIR / "prompts"
XLSX_PATH   = PROMPTS_DIR / "prompts_use.xlsx"
OUTPUT_DIR  = BASE_DIR / "output"


# ──────────────────────────────────────────────
# 日誌設定
# ──────────────────────────────────────────────
def setup_logging():
    log_format = "%(asctime)s [%(levelname)s] %(message)s"
    logging.basicConfig(
        level=logging.INFO,
        format=log_format,
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(BASE_DIR / "run.log", encoding="utf-8"),
        ],
    )


# ──────────────────────────────────────────────
# Excel 讀取
# ──────────────────────────────────────────────
def collect_excel_rows(xlsx_path: Path) -> list[dict]:
    """
    讀取 xlsx，回傳所有待處理的「列」資訊。
    每個 dict：
        sheet_name : 工作表名稱
        row_idx    : 列號（xlsx 的實際列號，從 2 開始）
        prompts    : [(seq, prompt_text), ...]  本列要依序處理的 prompt
    跳過 D 欄已有 Y 的列；跳過 B、C 都空的列。
    seq（流水號）在同一工作表內連續遞增，已完成的列也計入。
    """
    wb = openpyxl.load_workbook(xlsx_path)
    rows_to_process = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        seq = 0  # 每張工作表獨立計算

        for row_idx in range(2, ws.max_row + 1):
            b_val = ws.cell(row=row_idx, column=2).value
            c_val = ws.cell(row=row_idx, column=3).value
            d_val = ws.cell(row=row_idx, column=4).value

            b_text = str(b_val).strip() if b_val else ""
            c_text = str(c_val).strip() if c_val else ""

            # 計算本列的序號（不管是否已完成都要累計）
            row_prompts = []
            if b_text:
                seq += 1
                row_prompts.append((seq, b_text))
            if c_text:
                seq += 1
                row_prompts.append((seq, c_text))

            # 無內容 → 跳過
            if not row_prompts:
                continue

            # 已完成 → 計入 seq 但不加入待處理清單
            if d_val and str(d_val).strip().upper() == "Y":
                continue

            rows_to_process.append({
                "sheet_name": sheet_name,
                "row_idx":    row_idx,
                "prompts":    row_prompts,
            })

    return rows_to_process


def mark_row_done(xlsx_path: Path, sheet_name: str, row_idx: int):
    """在指定工作表、列的 D 欄寫入 Y 並存檔"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]
    ws.cell(row=row_idx, column=4).value = "Y"
    wb.save(xlsx_path)
    logging.getLogger(__name__).info(
        "已在 %s Row%d 標記完成（D 欄寫入 Y）", sheet_name, row_idx
    )


# ──────────────────────────────────────────────
# 輸出檔名
# ──────────────────────────────────────────────
def build_output_path(sheet_name: str, seq: int) -> Path:
    """
    output 檔名 = output/<日期>/<工作表名稱>_<流水號>_<日期>.mp4
    例：output/20260402/cute_pets_01_20260402.mp4
    """
    date_str = datetime.now().strftime("%Y%m%d")
    date_dir = OUTPUT_DIR / date_str
    date_dir.mkdir(parents=True, exist_ok=True)
    return date_dir / f"{sheet_name}_{seq:02d}_{date_str}.mp4"


# ──────────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────────
async def run():
    setup_logging()
    logger = logging.getLogger(__name__)

    # 載入 .env
    load_dotenv()
    username     = os.getenv("X_USERNAME", "").strip()
    password     = os.getenv("X_PASSWORD", "").strip()
    handle       = os.getenv("X_HANDLE", "").strip()
    login_method = os.getenv("LOGIN_METHOD", "x").strip().lower()

    if not username or not password:
        logger.error("請在 .env 中設定 X_USERNAME 與 X_PASSWORD")
        sys.exit(1)

    if login_method not in ("x", "email"):
        logger.warning("LOGIN_METHOD 值無效（%s），將使用預設值 'x'", login_method)
        login_method = "x"

    logger.info("登入方式：%s", login_method)

    # 確認 xlsx 存在
    if not XLSX_PATH.exists():
        logger.error("找不到 %s，請確認檔案路徑", XLSX_PATH)
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 讀取待處理列
    pending_rows = collect_excel_rows(XLSX_PATH)
    if not pending_rows:
        logger.info("xlsx 中沒有待處理的 prompt，程式結束")
        return

    total_prompts = sum(len(r["prompts"]) for r in pending_rows)
    logger.info(
        "共 %d 張工作表，%d 列待處理，%d 個 prompt，開始處理…",
        len({r["sheet_name"] for r in pending_rows}),
        len(pending_rows),
        total_prompts,
    )

    # 啟動自動化
    bot = GrokVideoAutomation(
        username=username,
        password=password,
        output_dir=OUTPUT_DIR,
        handle=handle,
        login_method=login_method,
    )
    await bot.start(headless=False)

    error_occurred = False
    success_count  = 0
    fail_count     = 0
    global_idx     = 0  # 跨所有 prompt 的進度計數

    try:
        # 登入（只需一次）
        if not await bot.ensure_logged_in():
            logger.error("登入失敗，程式中止")
            error_occurred = True
            return

        for row in pending_rows:
            sheet_name = row["sheet_name"]
            row_idx    = row["row_idx"]
            prompts    = row["prompts"]   # [(seq, text), ...]

            row_all_ok = True

            for seq, prompt_text in prompts:
                global_idx += 1
                output_path = build_output_path(sheet_name, seq)

                logger.info("=" * 60)
                logger.info(
                    "[%d/%d] 工作表：%s  Row%d  seq=%02d",
                    global_idx, total_prompts, sheet_name, row_idx, seq,
                )
                logger.info("Prompt：%s", prompt_text[:120])

                ok = await bot.generate_and_download(
                    prompt=prompt_text,
                    output_path=output_path,
                )

                if ok:
                    logger.info("成功：影片已儲存至 %s", output_path)
                    success_count += 1
                else:
                    logger.error(
                        "失敗：工作表 %s Row%d seq=%02d 未能生成影片",
                        sheet_name, row_idx, seq,
                    )
                    fail_count  += 1
                    row_all_ok   = False

                # 每個 prompt 之間稍作間隔
                if global_idx < total_prompts:
                    logger.info("等待 5 秒後處理下一個…")
                    await asyncio.sleep(5)

            # 整列所有 prompt 都成功 → 標記 Y
            if row_all_ok:
                mark_row_done(XLSX_PATH, sheet_name, row_idx)
            else:
                logger.warning(
                    "工作表 %s Row%d 有 prompt 失敗，不標記 Y，可重新執行補跑",
                    sheet_name, row_idx,
                )

    except Exception as e:
        logger.error("發生未預期錯誤：%s", e, exc_info=True)
        error_occurred = True

    finally:
        if error_occurred:
            logger.warning("=" * 60)
            logger.warning("發生錯誤，瀏覽器保持開啟以供檢查。")
            logger.warning("確認完畢後，請在此按 Enter 關閉瀏覽器…")
            await asyncio.get_event_loop().run_in_executor(None, input)
        await bot.stop()

    # 摘要
    logger.info("=" * 60)
    logger.info("全部完成：成功 %d 支 / 失敗 %d 支", success_count, fail_count)
    if fail_count > 0:
        logger.info("失敗的列未標記 Y，修正後可重新執行補跑")


if __name__ == "__main__":
    asyncio.run(run())
